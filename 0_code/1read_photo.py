# -*- coding: utf-8 -*-
"""
Created on Mon Dec 23 16:19:04 2024

@author: Lee
"""


from PIL import Image
from PIL.ExifTags import TAGS
import openpyxl
import os
import shutil
import pandas as pd
import json
import glob


def extract_exif(file_path) :
    '''
    이미지에서 촬영 기기와 화각 정보를 추출하는 코드
    '''
    
    try:
        image = Image.open(file_path)
        exif_data = image._getexif()
        if exif_data is None:
            return "Unknown", "Unknown"
        
        exif = {TAGS.get(tag): value for tag, value in exif_data.items() if tag in TAGS}
    
        #촬영 기기 정보가 없는 경우 Unknown으로 추출
        model = exif.get("Model","Unknown")
        
        #35mm 환산 화각 추출, 없는 경우 Unknown으로 추출
        focal_length = exif.get("FocalLengthIn35mmFilm","Unknown")
            
        return (model, focal_length)
    
    except Exception as e:
        return "Unknown", "Unknown"
        
def update_excel_with_exif(file_path, photo_directory):
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(file_path)
    
    # 모든 시트 순회
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # 첫 번째 행은 헤더, 2번째 행부터 시작
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=10):  # A부터 H까지 컬럼
            photo_path = row[7].value  # PHOTO PATH 컬럼 (8번째)
            model_cell = row[8]  # MODEL 컬럼 (9번째)
            focal_cell = row[9]  # FOCAL 컬럼 (10번째)
    
            if not photo_path:
                continue  # 사진 경로가 비어 있으면 건너뛰기
    
            # 사진 경로가 있고, 촬영 기기 또는 초점 거리가 공백인 경우 추출
            if (model_cell.value is None or model_cell.value == "") or (focal_cell.value is None or focal_cell.value == ""):
                if os.path.exists(photo_directory+photo_path):
                    model, focal_length = extract_exif(photo_directory+photo_path)

                    # 모델과 초점 거리 정보를 엑셀에 업데이트
                    if model_cell.value is None or model_cell.value.strip() == "":
                        model_cell.value = model 
                    if focal_cell.value is None or focal_cell.value.strip() == "":
                        focal_cell.value = focal_length
    
    # 엑셀 파일 저장
    wb.save(file_path)
    print("엑셀 파일이 업데이트되었습니다.")
    
    
#%%
do = input('데이터베이스를 업데이트 하시겠습니까?\n[y/n]\n>>> ')

if do.strip().lower()=='y':
    file_list = glob.glob('..\\*xlsx')
    if len(file_list) ==1 :
        file = file_list[0]
    else :
        print('데이터베이스로 추정되는 엑셀 파일 목록')
        for i in range(len(file_list)) :
            print('번호',i,':',file_list[i])
        file = int(input('사용할 데이터 파일 번호를 입력해주세요.\n>>> '))
        file = file_list[file]
    shutil.copy(file, file.replace('.xlsx','_bak.xlsx'))
    print('백업 파일이 생성되었습니다.')
    photo_dir = '..\\1_photo\\'
    
    update_excel_with_exif(file, photo_dir)
        
    ####json 만들기
    f = pd.ExcelFile(file)
    sheet_list = f.sheet_names
    f.close()
    
    for theater in sheet_list:
        if theater == '달오름극장':
            theater_data = pd.read_excel('..\\0_input\\THEATER_SEAT.xlsx', sheet_name=theater)
            name = 'dal'
        
        if theater == '하늘극장':
            theater_data = pd.read_excel('..\\0_input\\THEATER_SEAT.xlsx', sheet_name=theater)
            name = 'ha'
        
        if theater == '해오름극장':
            theater_data= pd.DataFrame()
            for floor in [1,2,3]:
                theater_data_floor = pd.read_excel('..\\0_input\\THEATER_SEAT.xlsx', sheet_name=theater+f'_{floor}층')
                theater_data = pd.concat([theater_data,theater_data_floor])
            name = 'hae'
                
        data = pd.read_excel(file, sheet_name=theater)
        
        data=data.merge(theater_data, on=['FLOOR', 'SECTION', 'ROW', 'NUMBER'])
        
        result=[]
        for i in data.index:
            if theater in ['달오름극장']:
                performance, date, floor, section, row, number, grade, photo_path, model, focal, right, ID, title, left, top, height, width = data.loc[i]
    
            if theater in ['해오름극장','하늘극장']:
                performance, date, floor, section, row, number, grade, photo_path, model, focal, right, ID, title, left, top, height, width, rotate = data.loc[i]
    
            # JSON 데이터 생성
            result.append({
                "id":photo_path,
                "seat_id": ID,
                "performanceName": performance,
                "performanceDate": date.strftime('%Y-%m-%d'),
                "seatNumber": f'{floor}층 {row}열 {number:02}번',
                "Grade" : grade,
                "image": '/1_photo/'+photo_path,
                "available": True,  # 기본값 True로 설정 (추가 로직 필요시 수정)
                "model": model,
                "focal": str(focal)+'mm',
                "copyright": right, 
            })
        # JSON 파일 저장
        with open(f'..\\0_input\\{name}.json', "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=4)
            
    print("json 파일이 업데이트되었습니다.")